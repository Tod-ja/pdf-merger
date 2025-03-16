from PyPDF2 import PdfWriter, PdfReader
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from io import BytesIO
from zipfile import ZipFile
from PIL import Image
import tempfile
import os
import subprocess
import img2pdf
from docx import Document
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import logging
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger('pdf_merger')

def convert_image_to_pdf(image_stream):
    try:
        logger.info("Converting image to PDF")
        image = Image.open(image_stream)

        # Use img2pdf for better quality conversion
        temp_image = tempfile.NamedTemporaryFile(delete=False, suffix='.png' if image.format == 'PNG' else '.jpg')
        image.save(temp_image, format=image.format)
        temp_image.close()

        temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        with open(temp_pdf.name, "wb") as f:
            f.write(img2pdf.convert(temp_image.name))

        # Read the PDF page
        temp_pdf.close()
        page = PdfReader(temp_pdf.name).pages[0]

        # Clean up
        os.remove(temp_image.name)
        os.remove(temp_pdf.name)

        return page
    except Exception as e:
        logger.error(f"Image conversion failed: {str(e)}")
        raise e

def convert_word_to_pdf(docx_stream):
    try:
        # Save the docx to a temporary file
        temp_docx = tempfile.NamedTemporaryFile(delete=False, suffix='.docx')
        docx_stream.seek(0)
        temp_docx.write(docx_stream.read())
        temp_docx.close()
        logger.info(f"Processing Word document: {temp_docx.name}")

        # Create output pdf file name
        temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        temp_pdf.close()

        conversion_success = False
        
        # Try LibreOffice on both Linux and Windows
        libreoffice_paths = [
            'libreoffice',  # Linux/Docker default
            'soffice',      # Alternative name on some systems
            r'C:\Program Files\LibreOffice\program\soffice.exe',  # Windows default path
            r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',  # Windows x86 path
        ]
        
        for libreoffice_path in libreoffice_paths:
            try:
                # Build the command based on OS
                cmd = [
                    libreoffice_path, 
                    '--headless', 
                    '--convert-to', 
                    'pdf', 
                    '--outdir', 
                    os.path.dirname(temp_pdf.name), 
                    temp_docx.name
                ]
                
                logger.info(f"Attempting Word conversion with LibreOffice: {libreoffice_path}")
                subprocess.run(cmd, check=True, capture_output=True)
                
                # Rename the output if needed (libreoffice creates a .pdf file with same basename)
                libreoffice_output = os.path.splitext(temp_docx.name)[0] + '.pdf'
                if os.path.exists(libreoffice_output) and libreoffice_output != temp_pdf.name:
                    os.replace(libreoffice_output, temp_pdf.name)
                
                logger.info(f"LibreOffice Word conversion successful")
                conversion_success = True
                break
            except (subprocess.CalledProcessError, FileNotFoundError) as e:
                logger.warning(f"LibreOffice conversion attempt failed: {str(e)}")
                continue
        
        # Use reportlab as final fallback
        if not conversion_success:
            try:
                # Fallback to generating a PDF from scratch with reportlab
                logger.info("Using fallback method (reportlab) for Word conversion")
                doc = Document(temp_docx.name)
                
                # Create a PDF
                c = canvas.Canvas(temp_pdf.name)
                y = 750  # Starting y position
                for para in doc.paragraphs:
                    text = para.text
                    if text:
                        c.drawString(72, y, text)
                        y -= 15  # Move down for next line
                        if y < 50:  # Create a new page if we run out of space
                            c.showPage()
                            y = 750
                c.save()
                conversion_success = True
                logger.info("Fallback Word conversion completed successfully")
            except Exception as e:
                logger.error(f"Fallback Word conversion failed: {str(e)}")
                raise Exception(f"All conversion methods failed. Last error: {str(e)}")

        # Clean up temporary docx file
        os.remove(temp_docx.name)

        # Return the PDF pages
        reader = PdfReader(temp_pdf.name)
        pages = reader.pages
        os.remove(temp_pdf.name)
        return pages
    except Exception as e:
        logger.error(f"Word conversion failed: {str(e)}")
        # Clean up any temp files if they exist
        try:
            if 'temp_docx' in locals() and os.path.exists(temp_docx.name):
                os.remove(temp_docx.name)
            if 'temp_pdf' in locals() and os.path.exists(temp_pdf.name):
                os.remove(temp_pdf.name)
        except:
            pass
        raise e

def convert_excel_to_pdf(excel_stream):
    try:
        # Save the Excel to a temporary file
        temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        excel_stream.seek(0)
        temp_excel.write(excel_stream.read())
        temp_excel.close()
        logger.info(f"Processing Excel document: {temp_excel.name}")

        # Create output pdf file name
        temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        temp_pdf.close()

        conversion_success = False
        
        # Try with LibreOffice using specific scale params for one page per sheet
        libreoffice_paths = [
            'libreoffice',  # Linux/Docker default
            'soffice',      # Alternative name on some systems
            r'C:\Program Files\LibreOffice\program\soffice.exe',  # Windows default path
            r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',  # Windows x86 path
        ]
        
        # First, pre-process the Excel file to set print settings
        try:
            logger.info("Pre-processing Excel file with direct print settings")
            wb = openpyxl.load_workbook(temp_excel.name)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
                    fitToPage=True
                )
                sheet.page_setup.fitToHeight = 1
                sheet.page_setup.fitToWidth = 1
                sheet.page_setup.orientation = 'landscape'
                sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
                
            # Save with print settings
            wb.save(temp_excel.name)
            logger.info("Successfully set print settings in Excel file")
        except Exception as e:
            logger.warning(f"Failed to pre-process Excel file: {str(e)}")
        
        # Try each possible LibreOffice path
        for libreoffice_path in libreoffice_paths:
            try:
                logger.info(f"Attempting Excel conversion with LibreOffice: {libreoffice_path}")
                cmd = [
                    libreoffice_path, 
                    '--headless', 
                    '--convert-to', 
                    'pdf:calc_pdf_Export:{"FitToPagesX":1,"FitToPagesY":1,"PageRange":"1-65535"}', 
                    '--outdir', 
                    os.path.dirname(temp_pdf.name), 
                    temp_excel.name
                ]
                
                subprocess.run(cmd, check=True, capture_output=True, timeout=60)
                
                # Rename the output if needed
                libreoffice_output = os.path.splitext(temp_excel.name)[0] + '.pdf'
                if os.path.exists(libreoffice_output) and libreoffice_output != temp_pdf.name:
                    os.replace(libreoffice_output, temp_pdf.name)
                
                logger.info("LibreOffice Excel conversion successful")
                conversion_success = True
                break
            except Exception as e:
                logger.warning(f"LibreOffice conversion attempt failed: {str(e)}")
                continue

        # Simple fallback method if LibreOffice fails
        if not conversion_success:
            logger.info("Using fallback method for Excel conversion")
            pdf_writer = PdfWriter()
            
            wb = openpyxl.load_workbook(temp_excel.name)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                buffer = BytesIO()
                c = canvas.Canvas(buffer, pagesize=(842, 595))  # A4 landscape
                
                # Title
                c.setFont("Helvetica-Bold", 16)
                c.drawString(30, 550, f"Sheet: {sheet_name}")
                
                # Draw table outline
                max_rows = min(sheet.max_row, 100)
                max_cols = min(sheet.max_column, 20)
                
                # Simple grid
                start_x, start_y = 30, 530
                cell_width, cell_height = 38, 20
                
                c.setFont("Helvetica", 8)
                for row in range(1, max_rows + 1):
                    for col in range(1, max_cols + 1):
                        x = start_x + (col - 1) * cell_width
                        y = start_y - (row) * cell_height
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value is not None:
                            c.drawString(x + 2, y + 5, str(cell_value)[:10])
                
                c.save()
                buffer.seek(0)
                pdf_reader = PdfReader(buffer)
                pdf_writer.add_page(pdf_reader.pages[0])
            
            with open(temp_pdf.name, 'wb') as f:
                pdf_writer.write(f)
            
            logger.info("Fallback Excel conversion completed")
            conversion_success = True

        # Clean up temporary files
        os.remove(temp_excel.name)

        # Return the PDF pages
        reader = PdfReader(temp_pdf.name)
        pages = reader.pages
        os.remove(temp_pdf.name)
        return pages
    except Exception as e:
        logger.error(f"Excel conversion failed: {str(e)}")
        # Clean up any temp files if they exist
        try:
            if 'temp_excel' in locals() and os.path.exists(temp_excel.name):
                os.remove(temp_excel.name)
            if 'temp_pdf' in locals() and os.path.exists(temp_pdf.name):
                os.remove(temp_pdf.name)
        except Exception as cleanup_err:
            logger.error(f"Error during cleanup: {str(cleanup_err)}")
        raise e

def convert_csv_to_pdf(csv_stream):
    try:
        # Save the CSV to a temporary file
        temp_csv = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
        csv_stream.seek(0)
        temp_csv.write(csv_stream.read())
        temp_csv.close()
        logger.info(f"Processing CSV document: {temp_csv.name}")

        # Create output pdf file name
        temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        temp_pdf.close()

        # First, try LibreOffice conversion since it handles formatting well
        conversion_success = False
        libreoffice_paths = [
            'libreoffice',  # Linux/Docker default
            'soffice',      # Alternative name on some systems
            r'C:\Program Files\LibreOffice\program\soffice.exe',  # Windows default path
            r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',  # Windows x86 path
        ]

        for libreoffice_path in libreoffice_paths:
            try:
                logger.info(f"Attempting CSV conversion with LibreOffice: {libreoffice_path}")
                cmd = [
                    libreoffice_path, 
                    '--headless', 
                    '--convert-to', 
                    'pdf:writer_pdf_Export', 
                    '--outdir', 
                    os.path.dirname(temp_pdf.name), 
                    temp_csv.name
                ]
                
                subprocess.run(cmd, check=True, capture_output=True, timeout=60)
                
                # Rename the output if needed
                libreoffice_output = os.path.splitext(temp_csv.name)[0] + '.pdf'
                if os.path.exists(libreoffice_output) and libreoffice_output != temp_pdf.name:
                    os.replace(libreoffice_output, temp_pdf.name)
                
                logger.info("LibreOffice CSV conversion successful")
                conversion_success = True
                break
            except Exception as e:
                logger.warning(f"LibreOffice conversion attempt failed: {str(e)}")
                continue

        # Fallback method using reportlab if LibreOffice fails
        if not conversion_success:
            logger.info("Using fallback method for CSV conversion")
            
            # Read CSV data
            import csv
            with open(temp_csv.name, 'r', newline='', encoding='utf-8') as f:
                csv_reader = csv.reader(f)
                data = list(csv_reader)

            if not data:
                logger.warning("CSV file is empty")
                # Create empty PDF
                c = canvas.Canvas(temp_pdf.name, pagesize=letter)
                c.drawString(72, 72, "Empty CSV file")
                c.save()
            else:
                # Style for table
                styles = getSampleStyleSheet()
                style_title = styles['Title']
                style_heading = styles['Heading1']
                style_normal = styles['Normal']

                # Determine data dimensions
                row_count = len(data)
                col_count = max(len(row) for row in data)

                # Prepare data for table
                table_data = []
                headers = data[0] if row_count > 0 else []
                
                # Ensure headers are complete (fill with empty strings if needed)
                headers = [h for h in headers] + [''] * (col_count - len(headers))
                table_data.append(headers)
                
                # Add rows
                for row in data[1:]:
                    # Ensure row has correct number of columns
                    row_data = [cell for cell in row] + [''] * (col_count - len(row))
                    table_data.append(row_data)

                # Create document
                doc = SimpleDocTemplate(temp_pdf.name, pagesize=landscape(letter))
                elements = []
                
                # Title
                csv_filename = os.path.basename(temp_csv.name)
                title = Paragraph(f"CSV File: {csv_filename}", style_title)
                elements.append(title)
                elements.append(Spacer(1, 12))
                
                # Set column widths - scale to page width
                page_width = letter[0] - 72  # Landscape width, minus margins
                col_width = page_width / col_count
                col_widths = [col_width] * col_count
                
                # Create table
                table = Table(table_data, colWidths=col_widths)
                
                # Add style to table
                table_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ])
                table.setStyle(table_style)
                
                elements.append(table)
                
                # Build document
                doc.build(elements)
            
            logger.info("Fallback CSV conversion completed successfully")
            conversion_success = True

        # Clean up temporary CSV file
        os.remove(temp_csv.name)

        # Return the PDF pages
        reader = PdfReader(temp_pdf.name)
        pages = reader.pages
        os.remove(temp_pdf.name)
        return pages
    
    except Exception as e:
        logger.error(f"CSV conversion failed: {str(e)}")
        # Clean up any temp files if they exist
        try:
            if 'temp_csv' in locals() and os.path.exists(temp_csv.name):
                os.remove(temp_csv.name)
            if 'temp_pdf' in locals() and os.path.exists(temp_pdf.name):
                os.remove(temp_pdf.name)
        except Exception as cleanup_err:
            logger.error(f"Error during cleanup: {str(cleanup_err)}")
        raise e

def add_label_to_page(page, label_text):
    width = float(page.mediabox.width)
    height = float(page.mediabox.height)
    rotation = page.get('/Rotate', 0)

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=(width, height))
    label_size = 60
    margin = 25

    # Determine label position based on rotation
    if rotation == 90:
        c.translate(height, 0)
        c.rotate(90)
        label_x = margin
        label_y = width - label_size - margin
    elif rotation == 180:
        c.translate(width, height)
        c.rotate(180)
        label_x = margin
        label_y = label_size + margin
    elif rotation == 270:
        c.translate(0, width)
        c.rotate(270)
        label_x = width - label_size - margin
        label_y = width - label_size - margin
    else:
        label_x = width - label_size - margin
        label_y = height - label_size - margin

    # Draw the label
    c.setFillColor(colors.transparent)
    c.setStrokeColor(colors.black)
    c.setLineWidth(2)
    c.circle(label_x + label_size / 2, label_y + label_size / 2, label_size / 2, stroke=1, fill=0)
    font_size = label_size / 2
    c.setFont("Helvetica-Bold", font_size)
    c.setFillColor(colors.black)
    text_width = c.stringWidth(label_text, "Helvetica-Bold", font_size)
    text_x = label_x + (label_size / 2) - (text_width / 2)
    text_y = label_y + (label_size / 2) - (font_size / 3)
    c.drawString(text_x, text_y, label_text)
    c.showPage()
    c.save()
    buffer.seek(0)
    label_page = PdfReader(buffer).pages[0]
    page.merge_page(label_page)

def merge_pdfs(files, labels, start_numbers):
    try:
        writer = PdfWriter()
        category_counts = {label: 0 for label in set(labels)}
        category_start_numbers = {label: start_num for label, start_num in zip(labels, start_numbers)}

        for file, label in zip(files, labels):
            file_stream = file.stream
            filename_lower = file.filename.lower()
            
            if filename_lower.endswith(('.jpg', '.jpeg', '.png')):
                file_stream.seek(0)
                page = convert_image_to_pdf(file_stream)
                pages = [page]
            elif filename_lower.endswith(('.doc', '.docx')):
                file_stream.seek(0)
                pages = convert_word_to_pdf(file_stream)
            elif filename_lower.endswith(('.xls', '.xlsx')):
                file_stream.seek(0)
                pages = convert_excel_to_pdf(file_stream)
            elif filename_lower.endswith(('.csv')):
                file_stream.seek(0)
                pages = convert_csv_to_pdf(file_stream)
            else:  # Assume PDF
                reader = PdfReader(file_stream)
                pages = reader.pages

            for page_num, page in enumerate(pages):
                if page_num == 0 and category_start_numbers[label] is not None:
                    current_number = category_start_numbers[label] + category_counts[label]
                    label_text = f"{label}{current_number}" if label.strip() else f"{current_number}"
                    category_counts[label] += 1
                    add_label_to_page(page, label_text)
                writer.add_page(page)

        output = BytesIO()
        writer.write(output)
        output.seek(0)
        return output
    except Exception as e:
        raise e

def label_pdfs(files, labels, start_numbers):
    try:
        category_files = {}
        for file, label in zip(files, labels):
            category_files.setdefault(label, []).append(file)

        labeled_files = []

        for label in category_files:
            files_in_category = category_files[label]
            start_number = start_numbers[labels.index(label)]

            for file in files_in_category:
                file_stream = file.stream
                original_filename = file.filename
                filename_lower = original_filename.lower()

                if filename_lower.endswith(('.jpg', '.jpeg', '.png')):
                    file_stream.seek(0)
                    page = convert_image_to_pdf(file_stream)
                    pages = [page]
                elif filename_lower.endswith(('.doc', '.docx')):
                    file_stream.seek(0)
                    pages = convert_word_to_pdf(file_stream)
                elif filename_lower.endswith(('.xls', '.xlsx')):
                    file_stream.seek(0)
                    pages = convert_excel_to_pdf(file_stream)
                elif filename_lower.endswith(('.csv')):
                    file_stream.seek(0)
                    pages = convert_csv_to_pdf(file_stream)
                else:  # Assume PDF
                    reader = PdfReader(file_stream)
                    pages = reader.pages

                writer = PdfWriter()
                for page_num, page in enumerate(pages):
                    if page_num == 0 and start_number is not None:
                        label_text = f"{label}{start_number}" if label.strip() else f"{start_number}"
                        start_number += 1
                        add_label_to_page(page, label_text)
                    writer.add_page(page)

                # Ensure output filename ends with .pdf
                output_filename = original_filename
                if not output_filename.lower().endswith('.pdf'):
                    base_name = os.path.splitext(original_filename)[0]
                    output_filename = f"{base_name}.pdf"

                temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
                writer.write(temp_pdf.name)
                temp_pdf.close()
                labeled_files.append((temp_pdf.name, output_filename))

        zip_buffer = BytesIO()
        with ZipFile(zip_buffer, 'w') as z:
            for temp_pdf_path, output_filename in labeled_files:
                z.write(temp_pdf_path, output_filename)
                os.remove(temp_pdf_path)
        zip_buffer.seek(0)
        return zip_buffer
    except Exception as e:
        raise e

def process_files(files, labels, output_filename, generate_cover=True):
    """
    Process multiple files, convert them to PDF if necessary, optionally add labels,
    and merge them into a single PDF.
    """
    logger.info(f"Processing {len(files)} files for merging")
    
    pdf_writer = PdfWriter()
    
    # Process each file and add to the output PDF
    for i, file in enumerate(files):
        file_ext = os.path.splitext(file.filename)[1].lower()
        label = labels[i] if labels and i < len(labels) else None
        
        logger.info(f"Processing file {i+1}/{len(files)}: {file.filename} (ext: {file_ext})")
        
        # Process different file types
        if file_ext in ['.pdf']:
            logger.info(f"File is already PDF, just reading")
            pages = PdfReader(BytesIO(file.read())).pages
        elif file_ext in ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.tif']:
            logger.info(f"Converting image file to PDF")
            pages = convert_image_to_pdf(BytesIO(file.read()))
        elif file_ext in ['.doc', '.docx']:
            logger.info(f"Converting Word document to PDF")
            pages = convert_word_to_pdf(BytesIO(file.read()))
        elif file_ext in ['.xls', '.xlsx']:
            logger.info(f"Converting Excel file to PDF")
            pages = convert_excel_to_pdf(BytesIO(file.read()))
        elif file_ext in ['.csv']:
            logger.info(f"Converting CSV file to PDF")
            pages = convert_csv_to_pdf(BytesIO(file.read()))
        else:
            logger.warning(f"Unsupported file type: {file_ext}")
            raise ValueError(f"Unsupported file type: {file_ext}")

        # Apply labels if provided
        if label:
            logger.info(f"Adding label '{label}' to file")
            for page in pages:
                add_label_to_page(page, label)
        
        # Add pages to PDF writer
        for page in pages:
            pdf_writer.add_page(page)

    # Generate a cover page if requested
    if generate_cover and labels:
        logger.info("Generating cover page with labels")
        cover_page = create_cover_page(labels)
        # Add the cover page as the first page
        pdf_writer.insert_page(cover_page, 0)
    
    # Write the output to a BytesIO object
    output_pdf = BytesIO()
    pdf_writer.write(output_pdf)
    output_pdf.seek(0)
    
    logger.info(f"PDF successfully merged into {output_filename}")
    return output_pdf

def create_cover_page(labels):
    # Create a new PDF with a single page
    logger.info(f"Creating cover page with {len(labels)} labels")
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=(595, 842))  # A4 size
    c.setFont("Helvetica-Bold", 24)
    c.drawString(72, 750, "Cover Page")
    c.setFont("Helvetica", 18)
    y = 700
    for label in labels:
        c.drawString(72, y, label)
        y -= 20
    c.showPage()
    c.save()
    buffer.seek(0)
    logger.info("Cover page created successfully")
    return PdfReader(buffer).pages[0]
